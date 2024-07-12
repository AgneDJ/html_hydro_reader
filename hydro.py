import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import glob
import os

# Reads percentages from HTML file


def read_html_table(html_file):
    with open(html_file, 'r') as file:
        soup = BeautifulSoup(file, 'html.parser')
    table = soup.find('table', {'id': 'table1'})
    if table:
        print(f"Table found in HTML file: {html_file}")
    else:
        print(f"No table found in HTML file: {html_file}")
        return []  # Return empty list if no table is found

    data = []
    for row in table.find_all('tr')[1:]:  # Skip the header row by slicing
        cols = row.find_all('td')
        cols = [ele.text.strip() for ele in cols]
        if cols:
            row_name = cols[0]  # First value is the row name
            # Remove percentage symbols and convert to float, skip the first column
            row_data = []
            for ele in cols[1:]:
                try:
                    value = float(ele.replace('%', ''))
                except ValueError:
                    value = None  # Set to None if conversion fails
                row_data.append(value)
            data.append((row_name, row_data))
    return data

# Unmerges cells in the given worksheet


def unmerge_cells(ws):
    merged_cells = list(ws.merged_cells)
    for merged_cell in merged_cells:
        ws.unmerge_cells(str(merged_cell))

# Get the column index based on the folder name


def get_column_index(ws, folder_name):
    for col in ws.iter_cols(min_row=7, max_row=7, min_col=4, max_col=ws.max_column):
        column_value = str(col[0].value)  # Convert column value to string
        print(
            f"Checking column {col[0].column} with value {column_value} against folder {folder_name}")
        if column_value == folder_name:
            return col[0].column
    return None

# Updates Excel file with percentages


def update_excel_with_percentages(excel_file, base_folder, sheet_name):
    # Get a list of all folders in the specified path
    folder_path = os.path.join(base_folder, '2024', 'Liepa', '*')
    folder_list = [f for f in glob.glob(folder_path) if os.path.isdir(f)]

    print(f"Found folders: {folder_list}")

    # Check if any folders were found
    if not folder_list:
        print("No folders found. Check the path and try again.")
        return

    # Loads the workbook and worksheet
    wb = load_workbook(excel_file)
    print("Available sheet names:", wb.sheetnames)
    if sheet_name not in wb.sheetnames:
        print(f"Worksheet {sheet_name} does not exist.")
        return

    ws = wb[sheet_name]

    # Unmerges cells
    unmerge_cells(ws)

    for folder in folder_list:
        folder_name = os.path.basename(folder)
        column_index = get_column_index(ws, folder_name)
        if column_index is None:
            print(f"No matching column for folder {folder_name}")
            continue

        # Get a list of all HTML files in the folder
        html_files = glob.glob(os.path.join(folder, '*.html'))

        for html_file in html_files:
            # Reads HTML table
            data = read_html_table(html_file)
            if not data:
                continue  # Skip if no data

            # Updates Excel sheet with percentages
            for row_name, values in data:
                # Column C is the 3rd column
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
                    if row[0].value == row_name:
                        row_idx = row[0].row
                        for j, value in enumerate(values):
                            if value is not None:  # Only write non-None values
                                ws.cell(row=row_idx,
                                        column=column_index + j, value=value)
                                print(
                                    f"Writing value {value} to cell ({row_idx}, {column_index + j})")

    # Saves the workbook
    wb.save(excel_file)
    print(f"Workbook {excel_file} saved successfully.")


# Base folder containing HTML files
base_folder = '/Users/djacenko/Desktop/html_failai/'

excel_file = '/Users/djacenko/Desktop/Pasitvirtinimo_vertinimas.xlsx'
sheet_name = 'Sheet1'  # Update the sheet name to 'Sheet1'

update_excel_with_percentages(excel_file, base_folder, sheet_name)
