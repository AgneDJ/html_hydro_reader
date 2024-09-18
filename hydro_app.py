import logging
from tkinter import ttk, filedialog, messagebox
import tkinter as tk
import os
import glob
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pandas as pd
from PIL import Image, ImageTk  # Import Pillow for image handling
import threading  # For running the update in a separate thread

# Set up logging
logging.basicConfig(filename='excel_updater.log', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def read_html_table(html_file):
    try:
        with open(html_file, 'r', encoding='utf-8') as file:
            soup = BeautifulSoup(file, 'html.parser')
    except Exception as e:
        logging.error(f"Error reading HTML file {html_file}: {e}")
        return []

    table = soup.find('table', {'id': 'table1'})
    if table:
        logging.info(f"Table found in HTML file: {html_file}")
    else:
        logging.warning(f"No table found in HTML file: {html_file}")
        return []  # Return empty list if no table is found

    data = []
    for row in table.find_all('tr')[1:]:  # Skip the header row by slicing
        cols = row.find_all('td')
        cols = [ele.text.strip() for ele in cols]
        if cols:
            row_name = cols[0]  # First value is the row name
            # Remove percentage symbols and convert to float, skip the first column
            row_data = []
            for ele in cols[1:]:
                try:
                    value = float(ele.replace('%', ''))
                except ValueError:
                    value = None  # Set to None if conversion fails
                row_data.append(value)
            data.append((row_name, row_data))
    return data


def unmerge_cells(ws):
    merged_cells = list(ws.merged_cells)
    for merged_cell in merged_cells:
        ws.unmerge_cells(str(merged_cell))


def get_column_index(ws, folder_name):
    for col in ws.iter_cols(min_row=7, max_row=7, min_col=4, max_col=ws.max_column):
        column_value = str(col[0].value)  # Convert column value to string
        logging.info(
            f"Checking column {col[0].column} with value {column_value} against folder {folder_name}")
        if column_value == folder_name:
            return col[0].column
    return None


def collect_data(base_folder, month_folder):
    collected_data = {}
    folder_path = os.path.join(base_folder, month_folder, '*')
    folder_list = [f for f in glob.glob(folder_path) if os.path.isdir(f)]

    logging.info(f"Found day folders: {folder_list}")

    for folder in folder_list:
        folder_name = os.path.basename(folder)
        html_files = glob.glob(os.path.join(folder, '*.html'))

        folder_data = []
        for html_file in html_files:
            # Reads HTML table
            data = read_html_table(html_file)
            if data:
                folder_data.append((folder_name, data))
        collected_data[folder_name] = folder_data

    return collected_data


def write_data_to_excel(excel_file, sheet_name, collected_data):
    try:
        # Load the workbook and worksheet
        wb = load_workbook(excel_file)
        logging.info("Available sheet names: " + ", ".join(wb.sheetnames))
        if sheet_name not in wb.sheetnames:
            logging.error(f"Worksheet {sheet_name} does not exist.")
            messagebox.showwarning(
                "Warning", f"Worksheet '{sheet_name}' does not exist.")
            return

        ws = wb[sheet_name]
        unmerge_cells(ws)

        total_items = sum(len(data) for folder_data in collected_data.values()
                          for _, data in folder_data)
        progress_increment = 100 / total_items if total_items > 0 else 0
        progress_value = 0

        for folder_name, data_entries in collected_data.items():
            column_index = get_column_index(ws, folder_name)
            if column_index is None:
                logging.warning(f"No matching column for folder {folder_name}")
                continue

            for _, data in data_entries:
                for row_name, values in data:
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
                        if row[0].value == row_name:
                            row_idx = row[0].row
                            for j, value in enumerate(values):
                                if value is not None:  # Only write non-None values
                                    ws.cell(
                                        row=row_idx, column=column_index + j, value=value)
                                    logging.info(
                                        f"Writing value {value} to cell ({row_idx}, {column_index + j})")
                            # Update progress
                            progress_value += progress_increment
                            progress_bar.step(progress_increment)
                            root.update_idletasks()

        # Save the workbook
        wb.save(excel_file)
        logging.info(f"Workbook {excel_file} saved successfully.")
        messagebox.showinfo("Success", "Data added successfully!")
    except Exception as e:
        logging.error(f"An error occurred while updating the Excel file: {e}")
        messagebox.showwarning("Warning", f"An error occurred: {e}")


def select_base_folder():
    year_folder_selected = filedialog.askdirectory(title="Select Year Folder")
    base_folder_entry.delete(0, tk.END)
    base_folder_entry.insert(0, year_folder_selected)

    # Populate month dropdown with subfolder names in the year folder
    if year_folder_selected:
        month_subfolders = [name for name in os.listdir(
            year_folder_selected) if os.path.isdir(os.path.join(year_folder_selected, name))]
        month_combo['values'] = month_subfolders
        month_combo.set('')  # Clear the current selection


def select_excel_file():
    file_selected = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")])
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_selected)

    # Populate sheet dropdown with sheet names
    if file_selected:
        wb = load_workbook(file_selected, read_only=True)
        sheets = wb.sheetnames
        sheet_name_combo['values'] = sheets
        sheet_name_combo.set('')  # Clear the current selection


def run_update():
    base_folder = base_folder_entry.get()
    month_folder = month_combo.get()
    excel_file = excel_file_entry.get()
    sheet_name = sheet_name_combo.get()

    if not base_folder or not excel_file or not sheet_name or not month_folder:
        messagebox.showerror("Error", "Please provide all inputs.")
        return

    # Reset the progress bar
    progress_bar['value'] = 0

    # Show the warning message
    warning_label.config(text="Adding data. Don't close the window.")
    warning_label.place(x=150, y=260)

    # Run the update process in a separate thread
    thread = threading.Thread(target=run_update_thread, args=(
        base_folder, month_folder, excel_file, sheet_name))
    thread.start()


def run_update_thread(base_folder, month_folder, excel_file, sheet_name):
    try:
        # Collect data from all HTML files first
        collected_data = collect_data(base_folder, month_folder)

        # Write collected data to the Excel file
        write_data_to_excel(excel_file, sheet_name, collected_data)
    finally:
        # Ensure the progress bar reaches 100% on success
        progress_bar['value'] = 100
        warning_label.config(text="")  # Clear the warning message
        root.update_idletasks()


def show_splash():
    global bg_image_splash
    splash = tk.Toplevel()
    splash.title("Welcome")
    splash.geometry("600x300")

    # Make the splash window always on top
    splash.attributes('-topmost', True)

    # Load and display the background image for the splash screen
    bg_image_splash = Image.open("images/splash_background.png")
    bg_image_splash = bg_image_splash.resize(
        (600, 300), Image.Resampling.LANCZOS)  # Resize to match window
    bg_photo_splash = ImageTk.PhotoImage(bg_image_splash)

    splash_label = tk.Label(splash, image=bg_photo_splash)
    # Keep a reference to avoid garbage collection
    splash_label.image = bg_photo_splash
    # Make the label fill the window
    splash_label.place(relwidth=1, relheight=1)

    # Add text on top of the image
    text_label = tk.Label(splash, text="Hydro Reader \n \n \n Author: Agne Djacenko", font=(
        "Helvetica", 8), bg='white')
    # Center the text label
    text_label.place(relx=0.5, rely=0.5, anchor='center')

    # Hide splash after 2 seconds and show the main window
    root.after(4000, splash.destroy)


def create_main_window():
    global bg_image_main, base_folder_entry, month_combo, excel_file_entry, sheet_name_combo, root, progress_bar, warning_label

    root.deiconify()  # Show the root window
    root.title("Hydro Reader")
    root.geometry("600x400")

    # Load and display the background image for the main window
    bg_image_main = Image.open("images/main_background.png")
    bg_image_main = bg_image_main.resize(
        (600, 400), Image.Resampling.LANCZOS)  # Resize to match window
    bg_photo_main = ImageTk.PhotoImage(bg_image_main)

    bg_label = tk.Label(root, image=bg_photo_main)
    bg_label.image = bg_photo_main  # Keep a reference to avoid garbage collection
    bg_label.place(relwidth=1, relheight=1)  # Make the label fill the window

    # Base folder selection
    tk.Label(root, text="Base Folder:").place(x=20, y=20)
    base_folder_entry = tk.Entry(root, width=50)
    base_folder_entry.place(x=100, y=20)
    tk.Button(root, text="Browse",
              command=select_base_folder).place(x=450, y=16)

    # Month selection dropdown
    tk.Label(root, text="Select Month:").place(x=20, y=60)
    month_combo = ttk.Combobox(root)
    month_combo.place(x=100, y=60)

    # Excel file selection
    tk.Label(root, text="Excel File:").place(x=20, y=100)
    excel_file_entry = tk.Entry(root, width=50)
    excel_file_entry.place(x=100, y=100)
    tk.Button(root, text="Browse", command=select_excel_file).place(x=450, y=96)

    # Sheet name dropdown
    tk.Label(root, text="Sheet Name:").place(x=20, y=140)
    sheet_name_combo = ttk.Combobox(root)
    sheet_name_combo.place(x=100, y=140)

    # Progress bar
    progress_bar = ttk.Progressbar(root, mode='determinate', length=300)
    progress_bar.place(x=150, y=220)

    # Warning message label
    warning_label = tk.Label(root, text="", fg="red", font=("Helvetica", 10))

    # Run button
    tk.Button(root, text="Run", command=run_update).place(x=260, y=180)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()  # Hide the root window for now
    show_splash()
    create_main_window()
    root.mainloop()
